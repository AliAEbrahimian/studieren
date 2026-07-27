# ============================================================
# IMPORTS
# ============================================================
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import SetPasswordForm
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth import get_user_model
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.core.mail import send_mail
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.conf import settings
from datetime import date
import openpyxl
import uuid

from .models import UserAccount
from academy.models import (
    Student, Employee, Course, Class, Enrollment,
    Session, Attendance, PlacementTestRequest,
    Exam, ExamSection, StudentGrade, OralGrade,
    PlacementTestSettings, ClassFeedback, WithdrawalRequest,
    TransferRequest, TaxSettings
)
from payments.models import Invoice
from .forms import RegisterForm, UserUpdateForm
from datetime import timedelta, datetime
from django.db.models import Sum, Count, Avg, Q, Max


User = get_user_model()


# ============================================================
# 1. AUTHENTICATION
# ============================================================

def loginPage(request):
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        email = request.POST.get('email', '').lower()
        password = request.POST.get('password', '')
        
        user = authenticate(request, username=email, password=password)
        
        if user is not None and user.is_active:
            try:
                emy = user.employee_profile
                messages.error(request, 'This login is for students only. Please use the staff login page.')
                return redirect('login')
            except Employee.DoesNotExist:
                pass
            
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid email or password.')
    
    context = {'page': 'login'}
    return render(request, 'base/login.html', context)


def staff_login(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        email = request.POST.get('email', '').lower()
        password = request.POST.get('password', '')
        
        user = authenticate(request, username=email, password=password)
        
        if user is not None and user.is_active:
            try:
                employee = user.employee_profile
                login(request, user)
                
                if employee.position in [Employee.Position.EXAM_MANAGER, Employee.Position.EXAM_CORRECTOR]:
                    return redirect('exam_dashboard')
                
                return redirect('dashboard')
            except Employee.DoesNotExist:
                messages.error(request, 'You do not have staff access. Please use the student login page.')
        else:
            messages.error(request, 'Invalid email or password.')
    
    context = {'page': 'staff_login'}
    return render(request, 'base/staff_login.html', context)


def logoutUser(request):
    logout(request)
    return redirect('login')


def registerPage(request):
    form = RegisterForm()
    
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.username = user.email.lower()
            user.is_active = False
            user.save()
            
            Student.objects.get_or_create(user=user, defaults={'enrollment_date': date.today()})
            
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            activation_link = request.build_absolute_uri(f'/activate/{uid}/{token}/')
            
            plain_message = (
                f'Hi {user.username},\n\n'
                f'Please click the link below to activate your account:\n'
                f'{activation_link}\n\n'
                f'If you did not sign up, ignore this message.'
            )
            
            html_message = f"""
                <p>Hi {user.username},</p>
                <p>Please click the link below to activate your account:</p>
                <p><a href="{activation_link}">Activate My Account</a></p>
                <p>If you did not sign up, ignore this message.</p>
            """
            
            send_mail(
                subject='Activate your account',
                message=plain_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[user.email],
                html_message=html_message,
                fail_silently=False,
            )
            
            messages.success(request, 'Please check your email to activate your account.')
            return redirect('login')
    
    return render(request, 'base/register.html', {'form': form})


def activateAccount(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    
    if user is not None and default_token_generator.check_token(user, token):
        user.is_active = True
        user.save()
        messages.success(request, 'Your account has been activated. You can now log in.')
        return redirect('login')
    else:
        messages.error(request, 'Invalid activation link.')
        return redirect('login')


@login_required(login_url='login')
def editProfile(request):
    user = request.user
    
    if request.method == 'POST':
        form = UserUpdateForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect('dashboard')
    else:
        form = UserUpdateForm(instance=user)
    
    context = {'form': form}
    return render(request, 'base/editprofile.html', context)


def resetPasswordRequest(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        
        try:
            user = User.objects.get(email=email)
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            reset_link = request.build_absolute_uri(f'/resetpassword/{uid}/{token}/')
            
            subject = 'Reset Password'
            message = (
                f"Hello {user.username},\n"
                f"We received a request to reset your password. To create a new one, please click the link below:\n\n"
                f"{reset_link}\n\n"
                f"If you didn't request this, you can safely ignore this message.\n"
                f"This link will expire in 1 hour.\n\n"
                f"Best regards,\n"
                f"The Studieren Team"
            )
            html_message = f"""
                <p>Hello {user.username},</p>
                <p>We received a request to reset your password.</p>
                <p>To create a new one, please click the link below:</p>
                <p><a href="{reset_link}">{reset_link}</a></p>
                <p>If you didn't request this, you can safely ignore this message.</p>
                <p>This link will expire in 1 hour.</p>
                <br>
                <p>Best regards,</p>
                <p>The Studieren Team</p>
            """
            
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[user.email],
                html_message=html_message,
                fail_silently=True,
            )
        except User.DoesNotExist:
            pass
        
        messages.success(request, 'If an account exists with this email, an email containing a recovery link will be sent to you.')
        return redirect('login')
    
    return render(request, 'base/resetpassword.html')


def resetPasswordConfirm(request, uidb64, token):
    request.session.flush()
    
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (User.DoesNotExist, ValueError, TypeError):
        user = None
    
    if user is not None and default_token_generator.check_token(user, token):
        form = SetPasswordForm(user, request.POST or None)
        
        if request.method == 'POST':
            if form.is_valid():
                form.save()
                messages.success(request, 'New password successfully set. Sign in now.')
                return redirect('login')
            else:
                messages.error(request, 'Password does not meet the required conditions.')
        
        context = {'form': form}
        return render(request, 'base/resetpassword_confirm.html', context)
    else:
        messages.error(request, 'The link is not valid or has expired.')
        return redirect('resetpassword')


# ============================================================
# 2. MAIN DASHBOARD
# ============================================================

@login_required(login_url='login')
def dashboard(request):
    user = request.user
    section = request.GET.get('section', 'overview')
    context = {
        'user': user,
        'total_revenue': 0,
        'today': date.today()
    }
    
    try:
        student_profile = user.student_profile
        context['is_student'] = True
        context['student'] = student_profile
        
        enrollments = Enrollment.objects.filter(
            student=student_profile
        ).select_related('enrolled_class', 'enrolled_class__course', 'enrolled_class__teacher__user')
        
        filter_type = request.GET.get('filter', 'current')
        today = date.today()
        if filter_type == 'finished':
            enrollments = enrollments.filter(enrolled_class__end_date__lt=today)
        else:
            enrollments = enrollments.filter(enrolled_class__end_date__gte=today)
        
        context['enrollments'] = enrollments
        context['current_filter'] = filter_type
        
        template_name = 'base/dashboard_student.html'
        
    except Student.DoesNotExist:
        context['is_student'] = False
    
    try:
        employee_profile = user.employee_profile
        context['is_employee'] = True
        context['employee'] = employee_profile
        
        if employee_profile.position == Employee.Position.TEACHER:
            today = date.today()
            taught_classes = employee_profile.taught_classes.filter(end_date__gte=today).select_related('course').order_by('-start_date')
            
            current_classes = taught_classes.filter(end_date__gte=today)
            finished_classes = taught_classes.filter(end_date__lt=today)
            unique_students = Student.objects.filter(
                enrollments__enrolled_class__teacher=employee_profile
            ).distinct().count()
            total_enrollments = sum(cls.enrollments.count() for cls in taught_classes)
            
            context['taught_classes'] = taught_classes
            context['current_classes_count'] = current_classes.count()
            context['finished_classes_count'] = finished_classes.count()
            context['total_classes_count'] = taught_classes.count()
            context['unique_students'] = unique_students
            context['total_enrollments'] = total_enrollments
            
            template_name = 'base/dashboard_teacher.html'
        
        elif employee_profile.position in [Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER]:
            total_students = Student.objects.count()
            total_teachers = Employee.objects.filter(position=Employee.Position.TEACHER).count()
            total_classes = Class.objects.filter(end_date__gte=date.today()).count()
            context.update({
                'total_students': total_students,
                'total_teachers': total_teachers,
                'total_classes': total_classes,
                'total_revenue': 0
            })
            
            context['is_senior_manager'] = (employee_profile.position == Employee.Position.SENIOR_MANAGER)
            
            template_name = 'base/dashboard_manager.html'
        
        else:
            template_name = 'base/dashboard_staff.html'
    
    except Employee.DoesNotExist:
        context['is_employee'] = False
        
        if not context.get('is_student'):
            template_name = 'base/dashboard_pending.html'
    
    return render(request, template_name, context)


# ============================================================
# 3. STUDENT PANEL
# ============================================================

@login_required(login_url='login')
def student_scores(request):
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, 'Only students can view scores.')
        return redirect('dashboard')
    
    # تمام ثبت‌نام‌های دانشجو
    enrollments = Enrollment.objects.filter(
        student=student
    ).select_related('enrolled_class__course', 'enrolled_class__teacher__user').order_by('-enrolled_class__end_date')
    
    scores_data = []
    for enrollment in enrollments:
        cls = enrollment.enrolled_class
        # فقط کلاس‌هایی که امتحان نهایی شده دارند
        if hasattr(cls, 'exam') and cls.exam.status == Exam.Status.FINALIZED:
            exam = cls.exam
            
            # نمرات کتبی
            written_grades = StudentGrade.objects.filter(
                student=student,
                exam_section__exam=exam
            ).select_related('exam_section')
            
            # نمره شفاهی
            oral_grade = OralGrade.objects.filter(student=student, exam=exam).first()
            
            total_written = sum(grade.score for grade in written_grades)
            oral_score = oral_grade.score if oral_grade else 0
            grand_total = total_written + oral_score
            
            # آیا بازخورد ثبت شده؟
            feedback_given = ClassFeedback.objects.filter(student=student, class_group=cls).exists()
            
            scores_data.append({
                'class': cls,
                'exam': exam,
                'written_grades': written_grades,
                'oral_grade': oral_grade,
                'total_written': total_written,
                'oral_score': oral_score,
                'grand_total': grand_total,
                'passed': grand_total >= exam.total_score * 0.6,  # ۶۰٪ نمره قبولی
                'feedback_given': feedback_given,
            })
    
    # خلاصهٔ وضعیت: آخرین سطح با نمره قبولی
    current_level = student.current_level
    latest_passed_level = None
    for item in scores_data:
        if item['passed']:
            latest_passed_level = item['class'].course.level
            break  # چون مرتب‌سازی از جدیدترین است
    
    
    passed_count = sum(1 for item in scores_data if item['passed'])
    failed_count = sum(1 for item in scores_data if item['passed'] is False)
    
    context = {
        'user': request.user,
        'student': student,
        'scores_data': scores_data,
        'latest_passed_level': latest_passed_level,
        'passed_count': passed_count,
        'failed_count': failed_count,
    }
    return render(request, 'base/student_scores.html', context)

@login_required(login_url='login')
def student_finance(request):
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, 'Only students can view financial records.')
        return redirect('dashboard')
    
    status_filter = request.GET.get('status', 'all')
    
    invoices = Invoice.objects.filter(
        student=student
    ).select_related('class_group__course')
    
    if status_filter in ['PAID', 'PENDING', 'CANCELED']:
        invoices = invoices.filter(status=status_filter)
    
    invoices = invoices.order_by('-created_at')
    
    total_amount = sum(inv.amount for inv in invoices)
    paid_amount = sum(inv.amount for inv in invoices if inv.status == Invoice.Status.PAID)
    pending_amount = sum(inv.amount for inv in invoices if inv.status == Invoice.Status.PENDING)
    
    context = {
        'user': request.user,
        'student': student,
        'invoices': invoices,
        'status_filter': status_filter,
        'total_amount': total_amount,
        'paid_amount': paid_amount,
        'pending_amount': pending_amount,
    }
    return render(request, 'base/student_finance.html', context)


@login_required(login_url='login')
def available_courses(request):
    today = date.today()
    
    courses = Class.objects.filter(
        end_date__gte=today
    ).select_related('course', 'teacher__user').order_by('start_date')
    
    search_query = request.GET.get('q', '')
    language_filter = request.GET.get('language', '')
    level_filter = request.GET.get('level', '')
    type_filter = request.GET.get('type', '')
    
    if search_query:
        courses = courses.filter(
            Q(title__icontains=search_query) |
            Q(course__title__icontains=search_query) |
            Q(course__language__icontains=search_query)
        )
    
    if language_filter:
        courses = courses.filter(course__language__iexact=language_filter)
    
    if level_filter:
        courses = courses.filter(course__level__iexact=level_filter)
    
    if type_filter:
        courses = courses.filter(class_type=type_filter)
    
    context = {
        'courses': courses,
        'search_query': search_query,
        'language_filter': language_filter,
        'level_filter': level_filter,
        'type_filter': type_filter,
    }
    return render(request, 'base/available_courses.html', context)


@login_required(login_url='login')
def class_detail(request, class_id):
    cls = get_object_or_404(Class, id=class_id)
    
    remaining_seats = cls.capacity - cls.enrollments.count()
    
    already_enrolled = False
    if hasattr(request.user, 'student_profile'):
        already_enrolled = Enrollment.objects.filter(
            student=request.user.student_profile,
            enrolled_class=cls
        ).exists()
        
    is_student = hasattr(request.user, 'student_profile')
    
    if is_student:
        already_enrolled = Enrollment.objects.filter(
            student=request.user.student_profile,
            enrolled_class=cls
        ).exists()
    
    context = {
        'class': cls,
        'remaining_seats': remaining_seats,
        'already_enrolled': already_enrolled,
        'is_student': is_student,
    }
    return render(request, 'base/class_detail.html', context)


@login_required(login_url='login')
def enroll_class(request, class_id):
    print("DEBUG: enroll_class view called!")
    cls = get_object_or_404(Class, id=class_id)
    
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, 'Only students can enroll in classes.')
        return redirect('available_courses')
    
    remaining_seats = cls.capacity - cls.enrollments.count()
    
    deadline = cls.start_date + timedelta(days=3)
    if date.today() > deadline:
        messages.error(request, f'Enrollment for this class closed on {deadline.strftime("%Y-%m-%d")}.')
        return redirect('class_detail', class_id=cls.id)
    
    if remaining_seats <= 0:
        print("DEBUG: Class is full!")
        messages.error(request, 'Sorry, this class is already full.')
        return redirect('class_detail', class_id=cls.id)
    
    if Enrollment.objects.filter(student=student, enrolled_class=cls).exists():
        print("DEBUG: Student is already enrolled!")
        messages.warning(request, 'You are already enrolled in this class.')
        return redirect('class_detail', class_id=cls.id)
    
    LEVEL_ORDER = {'A1': 1, 'A2': 2, 'B1': 3, 'B2': 4, 'C1': 5, 'C2': 6}
    
    student_level = student.current_level.upper() if student.current_level else ''
    class_level = cls.course.level
    
    if student_level and class_level:
        student_level_num = LEVEL_ORDER.get(student_level, 0)
        class_level_num = LEVEL_ORDER.get(class_level, 0)
        
        print(f"DEBUG: Student level: {student_level_num}, Class level: {class_level_num}")
        if student_level_num < class_level_num:
            messages.error(
                request,
                f'Your current level ({student_level}) is lower than the required level for this class ({class_level}). '
                f'Please complete lower levels first or request a placement test.'
            )
            return redirect('class_detail', class_id=cls.id)
    
    return redirect('enroll_review', class_id=cls.id)


@login_required(login_url='login')
def enroll_review(request, class_id):
    cls = get_object_or_404(Class, id=class_id)
    
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, 'Only students can enroll.')
        return redirect('available_courses')
    
    if Enrollment.objects.filter(student=student, enrolled_class=cls).exists():
        messages.warning(request, 'You are already enrolled.')
        return redirect('class_detail', class_id=cls.id)
    
    context = {
        'class': cls,
        'student': student,
    }
    return render(request, 'base/enroll_review.html', context)


@login_required(login_url='login')
def mock_payment(request, class_id):
    cls = get_object_or_404(Class, id=class_id)
    
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, 'Only students can enroll.')
        return redirect('available_courses')
    
    if Enrollment.objects.filter(student=student, enrolled_class=cls).exists():
        messages.warning(request, 'You are already enrolled.')
        return redirect('class_detail', class_id=cls.id)
    
    if request.method == 'POST':
        tax_settings = TaxSettings.load()
        tax_amount = (cls.tuition_fee * tax_settings.tax_percent) / 100
        
        Invoice.objects.create(
            student=student,
            class_group=cls,
            amount=cls.tuition_fee,
            tax_amount=tax_amount,
            status=Invoice.Status.PAID,
            paid_at=timezone.now(),
            reference_code=f"REF-{uuid.uuid4().hex[:8].upper()}"
        )
        
        Enrollment.objects.create(
            student=student,
            enrolled_class=cls,
            payment_status=Enrollment.PaymentStatus.PAID
        )
        
        messages.success(request, f'You have successfully enrolled in {cls.title}!')
        return redirect('dashboard')
    
    context = {
        'class': cls,
        'student': student,
        'course_fee': cls.tuition_fee,
        'registration_fee': 0,
        'discount': 0,
        'total_due': cls.tuition_fee,
    }
    return render(request, 'base/mock_payment.html', context)


@login_required(login_url='login')
def pay_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    if not hasattr(request.user, 'student_profile') or request.user.student_profile != invoice.student:
        messages.error(request, 'You do not have permission to pay this invoice.')
        return redirect('student_finance')
    
    if invoice.status != Invoice.Status.PENDING:
        messages.warning(request, 'This invoice is not pending payment.')
        return redirect('student_finance')
    
    invoice.status = Invoice.Status.PAID
    invoice.paid_at = timezone.now()
    invoice.save()
    
    enrollment = Enrollment.objects.filter(
        student=invoice.student,
        enrolled_class=invoice.class_group,
        payment_status=Enrollment.PaymentStatus.PENDING
    ).first()
    if enrollment:
        enrollment.payment_status = Enrollment.PaymentStatus.PAID
        enrollment.save()
    
    messages.success(request, f'Invoice {invoice.reference_code or invoice.id} has been paid successfully.')
    return redirect('student_finance')


@login_required(login_url='login')
def invoice_receipt(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    if not (
        hasattr(request.user, 'student_profile') and request.user.student_profile == invoice.student
    ) and not (
        request.user.is_staff or (
            hasattr(request.user, 'employee_profile') and
            request.user.employee_profile.position in [
                Employee.Position.STAFF, Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
            ]
        )
    ):
        messages.error(request, 'You do not have permission to view this receipt.')
        return redirect('dashboard')
    
    tax_settings = TaxSettings.load()
    total_amount = invoice.amount + invoice.tax_amount

    context = {
        'invoice': invoice,
        'tax_settings': tax_settings,
        'total_amount': total_amount,
    }
    return render(request, 'base/invoice_receipt.html', context)


@login_required(login_url='login')
def request_placement_test(request):
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, 'Only students can request a placement test.')
        return redirect('dashboard')
    
    open_request = PlacementTestRequest.objects.filter(student=student, status='PENDING').first()
    if open_request:
        messages.warning(request, 'You already have a pending placement test request.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        test_type = request.POST.get('test_type')
        requested_level = request.POST.get('requested_level', '')
        
        PlacementTestRequest.objects.create(
            student=student,
            test_type=test_type,
            requested_level=requested_level,
        )
        
        messages.success(request, 'Your placement test request has been submitted.')
        return redirect('dashboard')
    
    context = {
        'user': request.user,
        'student': student,
        'test_types': PlacementTestRequest.TestType.choices,
    }
    return render(request, 'base/request_placement_test.html', context)

@login_required(login_url='login')
def student_placement_status(request):
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, 'Only students can view placement status.')
        return redirect('dashboard')
    
    # آخرین درخواست (یا همه درخواست‌ها)
    requests = PlacementTestRequest.objects.filter(student=student).order_by('-created_at')
    settings = PlacementTestSettings.load()
    
    context = {
        'requests': requests,
        'settings': settings,
    }
    return render(request, 'base/student_placement_status.html', context)

@login_required(login_url='login')
def pay_placement_test(request, request_id):
    placement_request = get_object_or_404(PlacementTestRequest, id=request_id)
    
    # فقط دانشجوی صاحب درخواست
    if not hasattr(request.user, 'student_profile') or request.user.student_profile != placement_request.student:
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    if placement_request.status != PlacementTestRequest.Status.APPROVED:
        messages.error(request, 'Your request has not been approved yet.')
        return redirect('student_placement_status')
    
    settings = PlacementTestSettings.load()
    
    if request.method == 'POST':
        tax_settings = TaxSettings.load()
        tax_amount = (settings.test_fee * tax_settings.tax_percent) / 100
        
        # ایجاد فاکتور و پرداخت
        Invoice.objects.create(
            student=placement_request.student,
            class_group=None,  # برای کلاس نیست
            placement_request=placement_request,
            amount=settings.test_fee,
            tax_amount=tax_amount,
            status=Invoice.Status.PAID,
            paid_at=timezone.now(),
            reference_code=f"REF-PLACE-{uuid.uuid4().hex[:8].upper()}"
        )
        placement_request.payment_status = Enrollment.PaymentStatus.PAID
        placement_request.save()
        messages.success(request, 'Placement test fee paid successfully.')
        return redirect('student_placement_status')
    
    context = {
        'placement_request': placement_request,
        'test_fee': settings.test_fee,
    }
    return render(request, 'base/pay_placement_test.html', context)

@login_required(login_url='login')
def request_transfer(request, enrollment_id):
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, 'Only students can request transfer.')
        return redirect('dashboard')

    from_enrollment = get_object_or_404(Enrollment, id=enrollment_id, student=student)

    # اعتبارسنجی: فقط کلاس‌های فعال و با وضعیت ACTIVE
    if from_enrollment.status != Enrollment.EnrollmentStatus.ACTIVE:
        messages.error(request, 'This enrollment is not active.')
        return redirect('dashboard')
    if from_enrollment.enrolled_class.end_date < date.today():
        messages.error(request, 'This class has already ended.')
        return redirect('dashboard')

    # لیست کلاس‌های مقصد: کلاس‌های جاری که ظرفیت دارند و دانشجو قبلاً در آن ثبت‌نام نکرده
    available_classes = Class.objects.filter(
        end_date__gte=date.today()
    ).exclude(
        id__in=Enrollment.objects.filter(student=student).values_list('enrolled_class_id', flat=True)
    ).select_related('course', 'teacher__user').order_by('start_date')

    if request.method == 'POST':
        to_class_id = request.POST.get('to_class')
        reason = request.POST.get('reason', '')

        if not to_class_id:
            messages.error(request, 'Please select a destination class.')
            return redirect('request_transfer', enrollment_id=enrollment_id)

        to_class = get_object_or_404(Class, id=to_class_id)

        # چک تکراری نبودن درخواست
        if TransferRequest.objects.filter(
            student=student,
            from_enrollment=from_enrollment,
            to_class=to_class,
            status=TransferRequest.Status.PENDING
        ).exists():
            messages.warning(request, 'You already have a pending transfer request for this class.')
            return redirect('dashboard')

        TransferRequest.objects.create(
            student=student,
            from_enrollment=from_enrollment,
            to_class=to_class,
            reason=reason
        )
        messages.success(request, 'Transfer request submitted successfully. It is pending manager approval.')
        return redirect('dashboard')

    context = {
        'from_enrollment': from_enrollment,
        'available_classes': available_classes,
    }
    return render(request, 'base/request_transfer.html', context)


# ============================================================
# 4. TEACHER PANEL
# ============================================================

@login_required(login_url='login')
def teacher_classes(request):
    user = request.user
    try:
        employee = user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'You do not have a teacher profile.')
        return redirect('dashboard')
    
    all_classes = employee.taught_classes.all().select_related('course').order_by('-start_date')
    
    today = date.today()
    
    context = {
        'user': user,
        'employee': employee,
        'all_classes': all_classes,
        'today': today,
    }
    return render(request, 'base/teacher_classes.html', context)


@login_required(login_url='login')
def class_students(request, class_id):
    cls = get_object_or_404(Class, id=class_id)
    
    if request.user.employee_profile != cls.teacher:
        return redirect('dashboard')
    
    enrollments = Enrollment.objects.filter(enrolled_class=cls).select_related('student__user')
    context = {
        'cls': cls,
        'enrollments': enrollments,
    }
    return render(request, 'base/class_students.html', context)


@login_required(login_url='login')
def teacher_schedule(request, teacher_id=None):
    user = request.user
    
    if teacher_id and (request.user.is_staff or (
        hasattr(request.user, 'employee_profile') and
        request.user.employee_profile.position in [
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]
    )):
        teacher = get_object_or_404(Employee, pk=teacher_id, position=Employee.Position.TEACHER)
    else:
        try:
            teacher = request.user.employee_profile
            if teacher.position != Employee.Position.TEACHER:
                messages.error(request, 'You do not have a teacher profile.')
                return redirect('dashboard')
        except Employee.DoesNotExist:
            messages.error(request, 'You do not have a teacher profile.')
            return redirect('dashboard')
    
    today = date.today()
    classes = teacher.taught_classes.filter(end_date__gte=today).order_by('start_date')
    
    from collections import defaultdict
    schedule = defaultdict(list)
    for cls in classes:
        if cls.day_of_week:
            for day in cls.day_of_week:
                schedule[day].append(cls)
    
    for day in schedule:
        schedule[day].sort(key=lambda x: (x.start_time is None, x.start_time))
    
    context = {
        'user': user,
        'employee': teacher,
        'schedule': dict(schedule),
        'days': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'],
        'today': today,
        'viewing_as_manager': teacher_id is not None,
    }
    return render(request, 'base/teacher_schedule.html', context)


@login_required(login_url='login')
def teacher_attendance(request):
    user = request.user
    try:
        employee = user.employee_profile
    except Employee.DoesNotExist:
        messages.error(request, 'You do not have a teacher profile.')
        return redirect('dashboard')
    
    today = date.today()
    
    current_classes = employee.taught_classes.filter(
        end_date__gte=today
    ).order_by('start_date')
    
    context = {
        'user': user,
        'current_classes': current_classes,
    }
    return render(request, 'base/teacher_attendance.html', context)


@login_required(login_url='login')
def generate_class_sessions(request, class_id):
    cls = get_object_or_404(Class, id=class_id)
    
    if request.user.employee_profile != cls.teacher and not request.user.is_staff:
        messages.error(request, 'You do not have permission to perform this action.')
        return redirect('dashboard')
    
    count = cls.generate_sessions()
    messages.success(request, f'{count} new sessions were created successfully.')
    return redirect('attendance_sheet', class_id=class_id)


@login_required(login_url='login')
def attendance_sheet(request, class_id):
    cls = get_object_or_404(Class, id=class_id)
    
    is_teacher = hasattr(request.user, 'employee_profile') and request.user.employee_profile == cls.teacher
    is_manager = hasattr(request.user, 'employee_profile') and request.user.employee_profile.position in [
        Employee.Position.EDUCATION_MANAGER,
        Employee.Position.SENIOR_MANAGER
    ]
    if not (is_teacher or is_manager or request.user.is_staff):
        messages.error(request, 'You do not have permission to perform this action.')
        return redirect('dashboard')
    
    today = date.today()
    sessions = cls.sessions.filter(is_cancelled=False).order_by('date', 'start_time')
    enrollments = Enrollment.objects.filter(
        enrolled_class=cls,
    ).select_related('student__user').order_by('student__user__last_name')
    
    # یک دیکشنری کمکی: session_id -> session.date
    session_dates = {s.pk: s.date for s in sessions}
    
    attendance_data = {}
    for enrollment in enrollments:
        student = enrollment.student
        attendance_data[student.pk] = {}
        
        for session in sessions:
            if session.date > today:
                continue
            
            att = Attendance.objects.filter(session=session, student=student).first()
            attendance_data[student.pk][session.pk] = att  # اگر نباشد، None خواهد بود
    
    if request.method == 'POST':
        for key, value in request.POST.items():
            if key.startswith('status_'):
                parts = key.split('_')
                student_id = int(parts[1])
                session_id = int(parts[2])
                
                if session_id in session_dates and session_dates[session_id] > today:
                    continue
                
                Attendance.objects.update_or_create(
                    session_id=session_id,
                    student_id=student_id,
                    defaults={'status': value}
                )
        
        messages.success(request, 'Attendance saved successfully.')
        return redirect('attendance_sheet', class_id=class_id)
    
    context = {
        'cls': cls,
        'sessions': sessions,
        'enrollments': enrollments,
        'attendance_data': attendance_data,
        'today': today,
    }
    return render(request, 'base/attendance_sheet.html', context)


@login_required(login_url='login')
def export_attendance_excel(request, class_id):
    cls = get_object_or_404(Class, id=class_id)
    
    allowed_positions = [
        Employee.Position.STAFF,
        Employee.Position.EDUCATION_MANAGER,
        Employee.Position.SENIOR_MANAGER,
    ]
    is_teacher = hasattr(request.user, 'employee_profile') and request.user.employee_profile == cls.teacher
    is_allowed_staff = hasattr(request.user, 'employee_profile') and request.user.employee_profile.position in allowed_positions
    
    if not (is_teacher or is_allowed_staff or request.user.is_staff):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    sessions = cls.sessions.filter(is_cancelled=False).order_by('date', 'start_time')
    enrollments = Enrollment.objects.filter(
        enrolled_class=cls
    ).select_related('student__user').order_by('student__user__last_name')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    
    ws.cell(row=1, column=1, value="#")
    ws.cell(row=1, column=2, value="Student")
    for col, session in enumerate(sessions, start=3):
        ws.cell(row=1, column=col, value=session.date.strftime("%Y-%m-%d"))
    
    for row, enrollment in enumerate(enrollments, start=2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=enrollment.student.user.get_full_name())
        for col, session in enumerate(sessions, start=3):
            att = Attendance.objects.filter(session=session, student=enrollment.student).first()
            status = att.status if att else ''
            ws.cell(row=row, column=col, value=status)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=attendance_{cls.class_code}.xlsx'
    wb.save(response)
    return response

@login_required(login_url='login')
def teacher_performance(request):
    try:
        teacher = request.user.employee_profile
        if teacher.position != Employee.Position.TEACHER:
            messages.error(request, 'Only teachers can view performance.')
            return redirect('dashboard')
    except Employee.DoesNotExist:
        messages.error(request, 'You do not have a teacher profile.')
        return redirect('dashboard')
    
    # دریافت تمام کلاس‌هایی که این استاد تدریس کرده و بازخورد دارند
    classes_with_feedback = teacher.taught_classes.filter(
        feedbacks__isnull=False
    ).distinct()
    
    # محاسبه میانگین امتیازات
    feedbacks = ClassFeedback.objects.filter(class_group__teacher=teacher)
    agg = feedbacks.aggregate(
        avg_teaching=Avg('teaching_quality'),
        avg_communication=Avg('communication'),
        avg_punctuality=Avg('punctuality'),
        avg_engagement=Avg('engagement'),
        avg_overall=Avg('overall_satisfaction'),
        count=Count('id')
    )
    
    feedback_details = []
    recent_feedbacks = ClassFeedback.objects.filter(
        class_group__teacher=teacher
    ).select_related('student__user', 'class_group').order_by('-created_at')[:10]
    
    for fb in recent_feedbacks:
        feedback_details.append({
            'student_name': fb.student.user.get_full_name(),
            'class_title': fb.class_group.title,
            'teaching_quality': fb.teaching_quality,
            'communication': fb.communication,
            'punctuality': fb.punctuality,
            'engagement': fb.engagement,
            'overall': fb.overall_satisfaction,
            'created_at': fb.created_at,
        })
    
    context = {
        'teacher': teacher,
        'classes_with_feedback': classes_with_feedback,
        'avg_teaching': round(agg['avg_teaching'], 1) if agg['avg_teaching'] else 0,
        'avg_communication': round(agg['avg_communication'], 1) if agg['avg_communication'] else 0,
        'avg_punctuality': round(agg['avg_punctuality'], 1) if agg['avg_punctuality'] else 0,
        'avg_engagement': round(agg['avg_engagement'], 1) if agg['avg_engagement'] else 0,
        'avg_overall': round(agg['avg_overall'], 1) if agg['avg_overall'] else 0,
        'total_feedbacks': agg['count'],
        'feedback_details': feedback_details,
    }
    return render(request, 'base/teacher_performance.html', context)


# ============================================================
# 5. EXAM PANEL
# ============================================================

@login_required(login_url='login')
def enter_written_grades(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id)
    
    if not request.user.is_staff and request.user.employee_profile.position not in [
        Employee.Position.SENIOR_MANAGER,
        Employee.Position.EDUCATION_MANAGER,
        Employee.Position.EXAM_CORRECTOR,
        Employee.Position.EXAM_MANAGER
    ]:
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    sections = exam.sections.all().order_by('name')
    enrollments = Enrollment.objects.filter(enrolled_class=exam.class_group).select_related('student__user')
    
    if request.method == 'POST':
        for enrollment in enrollments:
            student = enrollment.student
            for section in sections:
                score = request.POST.get(f'score_{student.pk}_{section.pk}')
                if score:
                    StudentGrade.objects.update_or_create(
                        student=student,
                        exam_section=section,
                        defaults={'score': float(score), 'entered_by': request.user.employee_profile}
                    )
        messages.success(request, 'Written grades saved successfully.')
        return redirect('enter_written_grades', exam_id=exam.id)
    
    grades = {}
    for grade in StudentGrade.objects.filter(exam_section__exam=exam).select_related('student', 'exam_section'):
        student_pk = grade.student.pk
        section_pk = grade.exam_section.pk
        if student_pk not in grades:
            grades[student_pk] = {}
        grades[student_pk][section_pk] = grade.score
    
    oral_grades = {}
    for grade in OralGrade.objects.filter(exam=exam).select_related('student'):
        oral_grades[grade.student.pk] = grade.score
    
    context = {
        'exam': exam,
        'sections': sections,
        'enrollments': enrollments,
        'grades': grades,
        'oral_grades': oral_grades,
        'is_exam_manager': hasattr(request.user, 'employee_profile') and request.user.employee_profile.position in [
            Employee.Position.EXAM_MANAGER,
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ],
        
        'is_senior_manager': hasattr(request.user, 'employee_profile') and request.user.employee_profile.position == Employee.Position.SENIOR_MANAGER,
    }
    return render(request, 'base/enter_written_grades.html', context)


@login_required(login_url='login')
def enter_oral_grades(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id)
    cls = exam.class_group
    
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        (request.user.employee_profile != cls.teacher and
         request.user.employee_profile.position not in [
            Employee.Position.EXAM_MANAGER,
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ])
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    enrollments = Enrollment.objects.filter(
        enrolled_class=cls
    ).select_related('student__user').order_by('student__user__last_name')
    
    if request.method == 'POST':
        for enrollment in enrollments:
            student = enrollment.student
            score = request.POST.get(f'oral_score_{student.pk}')
            if score:
                OralGrade.objects.update_or_create(
                    student=student,
                    exam=exam,
                    defaults={
                        'score': float(score),
                        'entered_by': request.user.employee_profile if hasattr(request.user, 'employee_profile') else None
                    }
                )
        messages.success(request, 'Oral grades saved successfully.')
        return redirect('enter_oral_grades', exam_id=exam.id)
    
    oral_grades = {}
    for grade in OralGrade.objects.filter(exam=exam).select_related('student'):
        oral_grades[grade.student.pk] = grade.score
    
    context = {
        'exam': exam,
        'enrollments': enrollments,
        'oral_grades': oral_grades,
        'is_exam_manager': request.user.employee_profile.position == Employee.Position.EXAM_MANAGER if hasattr(request.user, 'employee_profile') else False,
    }
    return render(request, 'base/enter_oral_grades.html', context)


@login_required(login_url='login')
def exam_dashboard(request):
    user = request.user
    try:
        employee = user.employee_profile
        if employee.position not in [
            Employee.Position.EXAM_MANAGER,
            Employee.Position.EXAM_CORRECTOR,
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]:
            messages.error(request, 'You do not have permission to access the exam panel.')
            return redirect('dashboard')
    except Employee.DoesNotExist:
        messages.error(request, 'You do not have permission to access the exam panel.')
        return redirect('dashboard')
    
    today = date.today()
    context = {
        'user': user,
        'employee': employee,
        'is_exam_manager': employee.position in [
            Employee.Position.EXAM_MANAGER,
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ],
        'can_go_to_main_dashboard': employee.position in [
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]
    }
    
    open_exams = Exam.objects.filter(status=Exam.Status.OPEN).select_related('class_group__course').order_by('exam_date')
    context['open_exams'] = open_exams
    
    if employee.position in [
        Employee.Position.EXAM_MANAGER,
        Employee.Position.EDUCATION_MANAGER,
        Employee.Position.SENIOR_MANAGER
    ]:
        classes_without_exam = Class.objects.filter(
            end_date__lt=today
        ).exclude(
            id__in=Exam.objects.values_list('class_group_id', flat=True)
        ).select_related('course', 'teacher__user').order_by('-end_date')
        context['classes_without_exam'] = classes_without_exam
        
        finalized_exams = Exam.objects.filter(status=Exam.Status.FINALIZED).select_related('class_group__course').order_by('-finalized_at')
        context['finalized_exams'] = finalized_exams
        
        template_name = 'base/exam_manager_dashboard.html'
    else:
        template_name = 'base/exam_corrector_dashboard.html'
        # ==================== برای مصحح: آمار تصحیح‌های خودش ====================
        if employee.position == Employee.Position.EXAM_CORRECTOR:
            graded_exams_count = StudentGrade.objects.filter(
                entered_by=employee
            ).values('exam_section__exam').distinct().count()
    
            recent_exam_data = StudentGrade.objects.filter(
                entered_by=employee
            ).values('exam_section__exam').annotate(
                last_grade=Max('updated_at')
            ).order_by('-last_grade')[:5]
    
            exam_ids = [item['exam_section__exam'] for item in recent_exam_data]
            recent_graded_exams = Exam.objects.filter(
                id__in=exam_ids
            ).select_related('class_group__course')
    
            date_map = {item['exam_section__exam']: item['last_grade'] for item in recent_exam_data}
            for exam in recent_graded_exams:
                exam.graded_at = date_map.get(exam.id)
            recent_graded_exams = sorted(recent_graded_exams, key=lambda e: e.graded_at, reverse=True)
    
            context['graded_exams_count'] = graded_exams_count
            context['recent_graded_exams'] = recent_graded_exams
            
    return render(request, template_name, context)


@login_required(login_url='login')
def exam_list(request):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EXAM_MANAGER,
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    today = date.today()
    classes_without_exam = Class.objects.filter(
        end_date__lt=today
    ).exclude(
        id__in=Exam.objects.values_list('class_group_id', flat=True)
    ).select_related('course', 'teacher__user').order_by('-end_date')
    
    context = {
        'classes': classes_without_exam,
    }
    return render(request, 'base/exam_list.html', context)


@login_required(login_url='login')
def create_exam(request, class_id):
    cls = get_object_or_404(Class, id=class_id)
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        (request.user.employee_profile.position not in [
            Employee.Position.EXAM_MANAGER,
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ])
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    if hasattr(cls, 'exam'):
        messages.warning(request, 'An exam already exists for this class.')
        return redirect('exam_list')
    
    if request.method == 'POST':
        exam_date = request.POST.get('exam_date')
        total_score = request.POST.get('total_score', 300)
        exam = Exam.objects.create(
            class_group=cls,
            exam_date=exam_date,
            total_score=total_score,
        )
        section_names = request.POST.getlist('section_name[]')
        max_scores = request.POST.getlist('max_score[]')
        for name, max_s in zip(section_names, max_scores):
            if name.strip():
                ExamSection.objects.create(
                    exam=exam,
                    name=name.strip(),
                    max_score=max_s
                )
        messages.success(request, 'Exam created successfully with sections.')
        return redirect('exam_dashboard')
    
    context = {'class': cls}
    return render(request, 'base/create_exam.html', context)


@login_required(login_url='login')
def finalize_exam(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id)
    
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.SENIOR_MANAGER,
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.EXAM_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    if exam.status == Exam.Status.FINALIZED:
        messages.warning(request, 'Exam is already finalized.')
        return redirect('enter_written_grades', exam_id=exam.id)
    
    # چک کردن کامل بودن نمرات (فقط برای غیر از مدیر ارشد)
    if not request.user.employee_profile.position == Employee.Position.SENIOR_MANAGER:
        enrollments = exam.class_group.enrollments.all()
        missing_grades = []
        for enrollment in enrollments:
            student = enrollment.student
            # چک نمرات کتبی
            for section in exam.sections.all():
                if not StudentGrade.objects.filter(student=student, exam_section=section).exists():
                    missing_grades.append(f"{student.user.get_full_name()} - {section.name}")
            # چک نمره شفاهی
            if not OralGrade.objects.filter(student=student, exam=exam).exists():
                missing_grades.append(f"{student.user.get_full_name()} - Oral")
        
        if missing_grades:
            messages.error(request, f"All students must have complete written and oral grades before finalizing. Missing: {', '.join(missing_grades[:5])}{'...' if len(missing_grades) > 5 else ''}")
            return redirect('enter_written_grades', exam_id=exam.id)
    
    # نهایی‌سازی
    exam.status = Exam.Status.FINALIZED
    exam.finalized_by = request.user.employee_profile
    exam.finalized_at = timezone.now()
    exam.save()
    messages.success(request, 'Exam has been finalized. No further changes are allowed.')
    
    return redirect('enter_written_grades', exam_id=exam.id)

@login_required(login_url='login')
def reopen_exam(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id)
    
    # فقط مدیر ارشد یا ادمین
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position != Employee.Position.SENIOR_MANAGER
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    if exam.status == Exam.Status.OPEN:
        messages.warning(request, 'Exam is already open.')
        return redirect('enter_written_grades', exam_id=exam.id)
    
    exam.status = Exam.Status.OPEN
    exam.save()
    messages.success(request, 'Exam has been re-opened. Grades can be edited again.')
    return redirect('enter_written_grades', exam_id=exam.id)


@login_required(login_url='login')
def finalized_exam_list(request):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EXAM_MANAGER,
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    query = request.GET.get('q', '')
    exams = Exam.objects.filter(status=Exam.Status.FINALIZED).select_related('class_group__course').order_by('-finalized_at')
    
    if query:
        exams = exams.filter(
            Q(class_group__title__icontains=query) |
            Q(class_group__class_code__icontains=query) |
            Q(class_group__course__title__icontains=query)
        )
    
    context = {
        'exams': exams,
        'query': query,
    }
    return render(request, 'base/finalized_exam_list.html', context)


# ============================================================
# 6. STAFF PANEL
# ============================================================

@login_required(login_url='login')
def staff_enrollment(request):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.STAFF,
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    search_query = request.GET.get('q', '')
    selected_student = None
    student_results = []
    
    if search_query:
        student_results = UserAccount.objects.filter(
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(national_code__icontains=search_query),
            student_profile__isnull=False
        ).select_related('student_profile')[:10]
    
    student_id = request.GET.get('student_id', '')
    if student_id:
        selected_student = get_object_or_404(Student, pk=student_id)
    
    today = date.today()
    available_classes = Class.objects.filter(
        end_date__gte=today
    ).select_related('course', 'teacher__user').order_by('start_date')
    
    if request.method == 'POST':
        student_pk = request.POST.get('student_pk')
        class_pk = request.POST.get('class_pk')
        
        if student_pk and class_pk:
            student = get_object_or_404(Student, pk=student_pk)
            cls = get_object_or_404(Class, pk=class_pk)
            
            deadline = cls.start_date + timedelta(days=3)
            if date.today() > deadline:
                messages.error(request, f'Enrollment for {cls.title} closed on {deadline.strftime("%Y-%m-%d")}.')
                return redirect('staff_enrollment')
            
            if Enrollment.objects.filter(student=student, enrolled_class=cls).exists():
                messages.warning(request, f'{student.user.get_full_name()} is already enrolled in {cls.title}.')
            else:
                
                tax_settings = TaxSettings.load()
                tax_amount = (cls.tuition_fee * tax_settings.tax_percent) / 100
                
                Invoice.objects.create(
                    student=student,
                    class_group=cls,
                    amount=cls.tuition_fee,
                    tax_amount=tax_amount,
                    status=Invoice.Status.PAID,
                    paid_at=timezone.now(),
                    reference_code=f"REF-STAFF-{uuid.uuid4().hex[:8].upper()}"
                )
                Enrollment.objects.create(
                    student=student,
                    enrolled_class=cls,
                    payment_status=Enrollment.PaymentStatus.PAID
                )
                messages.success(request, f'{student.user.get_full_name()} has been enrolled in {cls.title}.')
                return redirect('staff_enrollment')
    
    context = {
        'search_query': search_query,
        'student_results': student_results,
        'selected_student': selected_student,
        'available_classes': available_classes,
        'total_students': Student.objects.count(),
        'available_classes_count': available_classes.count(),
    }
    return render(request, 'base/staff_enrollment.html', context)


@login_required(login_url='login')
def staff_student_profiles(request):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.STAFF,
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    search_query = request.GET.get('q', '')
    
    students = UserAccount.objects.filter(
        student_profile__isnull=False
    ).select_related('student_profile').order_by('last_name', 'first_name')
    
    if search_query:
        students = students.filter(
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(national_code__icontains=search_query)
        )
    
    today = date.today()
    
    student_data = []
    for user in students:
        student = user.student_profile
        active_enrollments = Enrollment.objects.filter(
            student=student,
            enrolled_class__end_date__gte=today
        ).count()
        student_data.append({
            'user': user,
            'student': student,
            'active_classes': active_enrollments,
        })
    
    context = {
        'student_data': student_data,
        'search_query': search_query,
    }
    return render(request, 'base/staff_student_profiles.html', context)


@login_required(login_url='login')
def staff_finance(request):
    # چک دسترسی: STAFF, EDUCATION_MANAGER, SENIOR_MANAGER
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.STAFF,
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    search_query = request.GET.get('q', '')
    status_filter = request.GET.get('status', 'all')
    
    # دریافت تمام فاکتورها
    invoices = Invoice.objects.select_related(
        'student__user', 'class_group__course'
    ).order_by('-created_at')
    
    # اعمال فیلتر وضعیت
    if status_filter in ['PAID', 'PENDING', 'CANCELED']:
        invoices = invoices.filter(status=status_filter)
    
    # اعمال جستجو
    if search_query:
        invoices = invoices.filter(
            Q(reference_code__icontains=search_query) |
            Q(student__user__first_name__icontains=search_query) |
            Q(student__user__last_name__icontains=search_query) |
            Q(student__user__email__icontains=search_query)
        )
    
    # آمار کلی
    total_revenue = invoices.filter(status=Invoice.Status.PAID).aggregate(Sum('amount'))['amount__sum'] or 0
    total_pending = invoices.filter(status=Invoice.Status.PENDING).aggregate(Sum('amount'))['amount__sum'] or 0
    paid_count = invoices.filter(status=Invoice.Status.PAID).count()
    pending_count = invoices.filter(status=Invoice.Status.PENDING).count()
    canceled_count = invoices.filter(status=Invoice.Status.CANCELED).count()
    
    context = {
        'user': request.user,
        'invoices': invoices,
        'search_query': search_query,
        'status_filter': status_filter,
        'total_revenue': total_revenue,
        'total_pending': total_pending,
        'paid_count': paid_count,
        'pending_count': pending_count,
        'canceled_count': canceled_count,
    }
    return render(request, 'base/staff_finance.html', context)

@login_required(login_url='login')
def student_detail(request, student_id):
    # ۱. چک دسترسی: STAFF, EDUCATION_MANAGER, SENIOR_MANAGER
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.STAFF,
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    # ۲. گرفتن دانشجو و اطلاعات کاربر
    student = get_object_or_404(Student.objects.select_related('user'), pk=student_id)
    user_profile = student.user

    # ۳. سوابق تحصیلی (Academic History)
    enrollments = Enrollment.objects.filter(
        student=student
    ).select_related('enrolled_class__course', 'enrolled_class__teacher__user').order_by('-enrolled_class__end_date')

    academic_history = []
    for enrollment in enrollments:
        cls = enrollment.enrolled_class
        passed = None
        if hasattr(cls, 'exam') and cls.exam.status == Exam.Status.FINALIZED:
            exam = cls.exam
            written_total = StudentGrade.objects.filter(
                student=student, exam_section__exam=exam
            ).aggregate(Sum('score'))['score__sum'] or 0
            oral = OralGrade.objects.filter(student=student, exam=exam).first()
            oral_score = oral.score if oral else 0
            passed = (written_total + oral_score) >= exam.total_score * 0.6
        
        academic_history.append({
            'class': cls,
            'course': cls.course,
            'teacher': cls.teacher,
            'registration_date': enrollment.registration_date,
            'end_date': cls.end_date,
            'is_active': cls.end_date >= date.today(),
            'passed': passed,
            'payment_status': enrollment.payment_status,
        })

    # ۴. وضعیت مالی
    invoices = Invoice.objects.filter(student=student).order_by('-created_at')
    total_paid = invoices.filter(status=Invoice.Status.PAID).aggregate(Sum('amount'))['amount__sum'] or 0
    total_pending = invoices.filter(status=Invoice.Status.PENDING).aggregate(Sum('amount'))['amount__sum'] or 0
    recent_invoices = invoices[:5]

    # ۵. آمار حضور و غیاب (حرفه‌ای و کامل)
    attendance_qs = Attendance.objects.filter(
        student=student,
        session__class_group__enrollments__student=student  # فقط جلسات کلاس‌هایی که واقعاً در آن ثبت‌نام کرده
    )
    total_sessions = attendance_qs.count()
    
    # محاسبه جزئیات وضعیت‌ها برای کارت‌های قالب
    present_count = attendance_qs.filter(status=Attendance.Status.PRESENT).count()
    absent_count = attendance_qs.filter(status=Attendance.Status.ABSENT).count()
    late_count = attendance_qs.filter(status=Attendance.Status.LATE).count()
    excused_count = attendance_qs.filter(status=Attendance.Status.EXCUSED).count()
    
    # درصد حضور
    if total_sessions > 0:
        attendance_percent = round((present_count / total_sessions * 100), 1)
    else:
        attendance_percent = 0

    # آخرین جلسات با وضعیت حضور
    recent_attendances = attendance_qs.select_related('session__class_group').order_by('-session__date')[:10]

    context = {
        'student': student,
        'user_profile': user_profile,
        'academic_history': academic_history,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'recent_invoices': recent_invoices,
        'attendance_percent': attendance_percent,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'excused_count': excused_count,
        'recent_attendances': recent_attendances,
    }
    return render(request, 'base/student_detail.html', context)

    # ۴. وضعیت مالی: خلاصه + آخرین ۵ فاکتور
    invoices = Invoice.objects.filter(student=student).order_by('-created_at')
    total_paid = invoices.filter(status=Invoice.Status.PAID).aggregate(Sum('amount'))['amount__sum'] or 0
    total_pending = invoices.filter(status=Invoice.Status.PENDING).aggregate(Sum('amount'))['amount__sum'] or 0
    recent_invoices = invoices[:5]

    # ۵. آمار حضور: درصد حضور کلی + آخرین جلسات
    total_sessions = Attendance.objects.filter(
        student=student, session__class_group__enrollments__student=student
    ).count()
    present_sessions = Attendance.objects.filter(
        student=student, status=Attendance.Status.PRESENT,
        session__class_group__enrollments__student=student
    ).count()
    attendance_percent = round((present_sessions / total_sessions * 100), 1) if total_sessions > 0 else 0

    # آخرین جلسات با وضعیت حضور
    recent_attendances = Attendance.objects.filter(
        student=student
    ).select_related('session__class_group').order_by('-session__date')[:10]

    context = {
        'student': student,
        'user_profile': user,
        'academic_history': academic_history,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'recent_invoices': recent_invoices,
        'attendance_percent': attendance_percent,
        'recent_attendances': recent_attendances,
    }
    return render(request, 'base/student_detail.html', context)

# ============================================================
# 7. MANAGER PANEL
# ============================================================

@login_required(login_url='login')
def manage_classes(request):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    search_query = request.GET.get('q', '')
    classes = Class.objects.select_related('course', 'teacher__user').order_by('-start_date')
    
    if search_query:
        classes = classes.filter(
            Q(title__icontains=search_query) |
            Q(class_code__icontains=search_query) |
            Q(course__language__icontains=search_query) |
            Q(course__level__icontains=search_query) |
            Q(teacher__user__first_name__icontains=search_query) |
            Q(teacher__user__last_name__icontains=search_query)
        )
    
    today = date.today()
    
    total_classes = classes.count()
    active_classes_count = classes.filter(end_date__gte=today).count()
    finished_classes_count = classes.filter(end_date__lt=today).count()
    
    context = {
        'classes': classes,
        'search_query': search_query,
        'today': today,
        'total_classes': total_classes,
        'active_classes_count': active_classes_count,
        'finished_classes_count': finished_classes_count,
    }
    return render(request, 'base/manage_classes.html', context)


@login_required(login_url='login')
def manage_courses(request):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    search_query = request.GET.get('q', '')
    courses = Course.objects.all().order_by('language', 'level')

    if search_query:
        courses = courses.filter(
            Q(title__icontains=search_query) |
            Q(language__icontains=search_query) |
            Q(level__icontains=search_query)
        )

    context = {
        'courses': courses,
        'search_query': search_query,
    }
    return render(request, 'base/manage_courses.html', context)


@login_required(login_url='login')
def create_course(request):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        language = request.POST.get('language', '').strip()
        level = request.POST.get('level', '').strip()
        description = request.POST.get('description', '').strip()

        if not title or not language or not level:
            messages.error(request, 'Please fill in all required fields.')
            return redirect('create_course')

        Course.objects.create(
            title=title, language=language, level=level, description=description
        )
        messages.success(request, f'Course "{title}" created successfully.')
        return redirect('manage_courses')

    context = {
        'levels': ['A1', 'A2', 'B1', 'B2', 'C1', 'C2'],
    }
    return render(request, 'base/create_course.html', context)


@login_required(login_url='login')
def edit_course(request, course_id):
    # دسترسی: مدیر آموزشی، مدیر ارشد
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    course = get_object_or_404(Course, pk=course_id)

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        language = request.POST.get('language', '').strip()
        level = request.POST.get('level', '').strip()
        description = request.POST.get('description', '').strip()

        if not title or not language or not level:
            messages.error(request, 'Please fill in all required fields.')
            return redirect('edit_course', course_id=course.id)

        # به‌روزرسانی فیلدها
        course.title = title
        course.language = language
        course.level = level
        course.description = description
        course.save()

        messages.success(request, f'Course "{course.title}" updated successfully.')
        return redirect('manage_courses')

    # GET: نمایش فرم با داده‌های فعلی
    context = {
        'course': course,
        'levels': ['A1', 'A2', 'B1', 'B2', 'C1', 'C2'],
    }
    return render(request, 'base/edit_course.html', context)


@login_required(login_url='login')
def delete_course(request, course_id):
    # دسترسی: مدیر آموزشی، مدیر ارشد
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    course = get_object_or_404(Course, pk=course_id)
    
    # بررسی اینکه آیا کلاسی به این دوره متصل است
    if course.classes.exists():
        messages.error(request, f'Cannot delete "{course.title}" because it has associated classes.')
        return redirect('manage_courses')
    
    course.delete()
    messages.success(request, f'Course "{course.title}" has been deleted.')
    return redirect('manage_courses')

@login_required(login_url='login')
def manage_teachers(request):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    search_query = request.GET.get('q', '')
    teachers = Employee.objects.filter(
        position=Employee.Position.TEACHER
    ).select_related('user').order_by('user__last_name', 'user__first_name')
    
    if search_query:
        teachers = teachers.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(department__icontains=search_query)
        )
    
    today = date.today()
    teacher_data = []
    for teacher in teachers:
        current_classes = teacher.taught_classes.filter(end_date__gte=today)
        total_classes = teacher.taught_classes.count()
        total_students = Student.objects.filter(
            enrollments__enrolled_class__teacher=teacher
        ).distinct().count()
        teacher_data.append({
            'teacher': teacher,
            'current_classes_count': current_classes.count(),
            'total_classes': total_classes,
            'total_students': total_students,
        })
    
    context = {
        'teacher_data': teacher_data,
        'search_query': search_query,
    }
    return render(request, 'base/manage_teachers.html', context)


@login_required(login_url='login')
def create_teacher(request):
    # دسترسی: مثل manage_teachers
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip().lower()
        password = request.POST.get('password', '')
        
        if not all([first_name, last_name, email, password]):
            messages.error(request, 'All fields are required.')
            return redirect('create_teacher')
        
        if UserAccount.objects.filter(email=email).exists():
            messages.error(request, 'A user with this email already exists.')
            return redirect('create_teacher')
        
        user = UserAccount.objects.create_user(
            email=email, password=password,
            first_name=first_name, last_name=last_name,
            username=email, is_active=True
        )
        Employee.objects.create(user=user, position=Employee.Position.TEACHER)
        
        messages.success(request, f'Teacher {user.get_full_name()} created successfully.')
        return redirect('manage_teachers')
    
    return render(request, 'base/create_teacher.html')


@login_required(login_url='login')
def edit_teacher(request, teacher_id):
    # دسترسی: مدیر آموزشی، مدیر ارشد
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    teacher = get_object_or_404(Employee, pk=teacher_id, position=Employee.Position.TEACHER)
    user = teacher.user

    # لیست سوپروایزرهای ممکن (همه کارمندان به جز خودش)
    supervisors = Employee.objects.exclude(pk=teacher.pk).select_related('user').order_by('user__last_name')

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        department = request.POST.get('department', '').strip()
        hire_date_str = request.POST.get('hire_date', '')
        supervisor_id = request.POST.get('supervisor', '')
        is_active = request.POST.get('is_active', 'on') == 'on'

        if not first_name or not last_name:
            messages.error(request, 'First name and last name are required.')
            return redirect('edit_teacher', teacher_id=teacher.pk)

        # به‌روزرسانی User
        user.first_name = first_name
        user.last_name = last_name
        user.is_active = is_active
        user.save()

        # به‌روزرسانی Employee
        teacher.department = department
        if hire_date_str:
            teacher.hire_date = date.fromisoformat(hire_date_str)
        else:
            teacher.hire_date = None
        
        if supervisor_id:
            teacher.supervisor = get_object_or_404(Employee, pk=supervisor_id)
        else:
            teacher.supervisor = None
        
        teacher.save()

        messages.success(request, f'Teacher {user.get_full_name()} updated successfully.')
        return redirect('manage_teachers')

    context = {
        'teacher': teacher,
        'supervisors': supervisors,
    }
    return render(request, 'base/edit_teacher.html', context)


@login_required(login_url='login')
def teacher_profile(request, teacher_id):
    teacher = get_object_or_404(Employee, pk=teacher_id, position=Employee.Position.TEACHER)
    # دسترسی
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER, Employee.Position.STAFF
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    today = date.today()
    classes = teacher.taught_classes.select_related('course').order_by('-start_date')
    current_classes = classes.filter(end_date__gte=today)
    finished_classes = classes.filter(end_date__lt=today)
    
    context = {
        'teacher': teacher,
        'current_classes': current_classes,
        'finished_classes': finished_classes,
    }
    return render(request, 'base/teacher_profile.html', context)


@login_required(login_url='login')
def teacher_assign_classes(request, teacher_id):
    teacher = get_object_or_404(Employee, pk=teacher_id, position=Employee.Position.TEACHER)
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    # کلاس‌هایی که این استاد تدریس می‌کند
    assigned_classes = teacher.taught_classes.select_related('course').order_by('start_date')
    # کلاس‌هایی که می‌توان به او اختصاص داد (هنوز معلم ندارند یا معلم دیگری دارند)
    today = date.today()
    available_classes = Class.objects.filter(end_date__gte=today).exclude(
        id__in=assigned_classes.values_list('id', flat=True)
    ).select_related('course').order_by('start_date')
    
    if request.method == 'POST':
        class_id = request.POST.get('class_id')
        action = request.POST.get('action')  # 'assign' or 'remove'
        cls = get_object_or_404(Class, pk=class_id)
        
        if action == 'assign':
            cls.teacher = teacher
            cls.save()
            messages.success(request, f'{teacher.user.get_full_name()} assigned to {cls.title}.')
        elif action == 'remove':
            if cls.teacher == teacher:
                cls.teacher = None
                cls.save()
                messages.success(request, f'{teacher.user.get_full_name()} removed from {cls.title}.')
        
        return redirect('teacher_assign_classes', teacher_id=teacher.pk)
    
    context = {
        'teacher': teacher,
        'assigned_classes': assigned_classes,
        'available_classes': available_classes,
    }
    return render(request, 'base/teacher_assign_classes.html', context)


@login_required(login_url='login')
def manager_reports(request):
    # ==================== ۱. دسترسی ====================
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    # ==================== ۲. خواندن فیلترها ====================
    today = date.today()
    date_range = request.GET.get('range', '30d')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    language_filter = request.GET.get('language', '')
    level_filter = request.GET.get('level', '')

    if date_range == '7d':
        start_date = today - timedelta(days=7)
        end_date = today
    elif date_range == '90d':
        start_date = today - timedelta(days=90)
        end_date = today
    elif date_range == 'custom' and start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        start_date = today - timedelta(days=30)
        end_date = today

    # ==================== ۳. فیلترهای پایه ====================
    invoice_filter = Q(status=Invoice.Status.PAID, paid_at__date__gte=start_date, paid_at__date__lte=end_date)
    enrollment_filter = Q(registration_date__gte=start_date, registration_date__lte=end_date)
    attendance_filter = Q(session__date__gte=start_date, session__date__lte=end_date)

    if language_filter:
        invoice_filter &= Q(class_group__course__language__iexact=language_filter)
    if level_filter:
        invoice_filter &= Q(class_group__course__level__iexact=level_filter)

    enrollment_filter &= Q(enrolled_class__course__language__iexact=language_filter) if language_filter else Q()
    enrollment_filter &= Q(enrolled_class__course__level__iexact=level_filter) if level_filter else Q()

    # ==================== ۴. کارت‌های KPI ====================
    total_revenue = Invoice.objects.filter(invoice_filter).aggregate(Sum('amount'))['amount__sum'] or 0
    active_students = Enrollment.objects.filter(enrolled_class__end_date__gte=today).values('student').distinct().count()
    active_classes_qs = Class.objects.filter(end_date__gte=today)
    if language_filter:
        active_classes_qs = active_classes_qs.filter(course__language__iexact=language_filter)
    if level_filter:
        active_classes_qs = active_classes_qs.filter(course__level__iexact=level_filter)
    active_classes_count = active_classes_qs.count()

    # کل دانشجویان و کلاس‌ها (بدون فیلتر تاریخ)
    total_students_all = Student.objects.count()
    total_classes_all = Class.objects.count()

    # حضور
    total_att = Attendance.objects.filter(attendance_filter).count()
    present_att = Attendance.objects.filter(attendance_filter, status=Attendance.Status.PRESENT).count()
    attendance_percent = round((present_att / total_att * 100), 1) if total_att > 0 else 0

    # نرخ تکمیل
    finished_classes_qs = Class.objects.filter(end_date__lt=today)
    if language_filter:
        finished_classes_qs = finished_classes_qs.filter(course__language__iexact=language_filter)
    if level_filter:
        finished_classes_qs = finished_classes_qs.filter(course__level__iexact=level_filter)
    finished_count = finished_classes_qs.count()
    completed_count = finished_classes_qs.filter(exam__status=Exam.Status.FINALIZED).count()
    completion_rate = round((completed_count / finished_count * 100), 1) if finished_count > 0 else 0

    # میانگین نمرات
    exam_grades = StudentGrade.objects.filter(
        exam_section__exam__status=Exam.Status.FINALIZED,
        exam_section__exam__class_group__end_date__gte=start_date,
        exam_section__exam__class_group__end_date__lte=end_date
    )
    if language_filter:
        exam_grades = exam_grades.filter(exam_section__exam__class_group__course__language__iexact=language_filter)
    if level_filter:
        exam_grades = exam_grades.filter(exam_section__exam__class_group__course__level__iexact=level_filter)
    avg_score = exam_grades.aggregate(Avg('score'))['score__avg'] or 0

    # نرخ قبولی کلی
    passed_students = 0
    total_exam_students = 0
    for exam in Exam.objects.filter(status=Exam.Status.FINALIZED):
        total = exam.total_score
        enrollments = exam.class_group.enrollments.all()
        for enrollment in enrollments:
            student = enrollment.student
            written_total = StudentGrade.objects.filter(student=student, exam_section__exam=exam).aggregate(Sum('score'))['score__sum'] or 0
            oral = OralGrade.objects.filter(student=student, exam=exam).first()
            oral_score = oral.score if oral else 0
            if (written_total + oral_score) >= total * 0.6:
                passed_students += 1
            total_exam_students += 1
    overall_pass_rate = round((passed_students / total_exam_students * 100), 1) if total_exam_students > 0 else 0

    # ==================== ۵. نمودارهای روند ====================
    # درآمد ماهانه
    monthly_revenue_labels = []
    monthly_revenue_data = []
    for i in range(5, -1, -1):
        month_start = today.replace(day=1) - timedelta(days=i*30)
        month_start = month_start.replace(day=1)
        if i == 0:
            month_end = today
        else:
            next_month = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_month - timedelta(days=next_month.day)
        month_q = invoice_filter & Q(paid_at__date__gte=month_start, paid_at__date__lte=month_end)
        month_revenue = Invoice.objects.filter(month_q).aggregate(Sum('amount'))['amount__sum'] or 0
        monthly_revenue_labels.append(month_start.strftime("%b %Y"))
        monthly_revenue_data.append(month_revenue)

    # ثبت‌نام ماهانه
    enrollment_labels = []
    enrollment_data = []
    for i in range(5, -1, -1):
        month_start = today.replace(day=1) - timedelta(days=i*30)
        month_start = month_start.replace(day=1)
        if i == 0:
            month_end = today
        else:
            next_month = month_start.replace(day=28) + timedelta(days=4)
            month_end = next_month - timedelta(days=next_month.day)
        month_q = enrollment_filter & Q(registration_date__date__gte=month_start, registration_date__date__lte=month_end)
        count = Enrollment.objects.filter(month_q).count()
        enrollment_labels.append(month_start.strftime("%b %Y"))
        enrollment_data.append(count)

    # حضور هفتگی
    weekly_att_labels = []
    weekly_att_data = []
    for i in range(3, -1, -1):
        week_end = today - timedelta(days=i*7)
        week_start = week_end - timedelta(days=6)
        week_q = attendance_filter & Q(session__date__gte=week_start, session__date__lte=week_end)
        week_total = Attendance.objects.filter(week_q).count()
        week_present = Attendance.objects.filter(week_q, status=Attendance.Status.PRESENT).count()
        percent = round((week_present / week_total * 100), 1) if week_total > 0 else 0
        weekly_att_labels.append(f"{week_start.strftime('%d %b')}")
        weekly_att_data.append(percent)

    # ==================== ۶. نمودارهای توزیع ====================
    # توزیع درآمد بر اساس زبان
    revenue_by_language = Invoice.objects.filter(invoice_filter).values(
        'class_group__course__language'
    ).annotate(total=Sum('amount')).order_by('-total')
    pie_language_labels = [item['class_group__course__language'] for item in revenue_by_language]
    pie_language_data = [float(item['total']) for item in revenue_by_language]

    # توزیع سطوح
    level_distribution = Student.objects.values('current_level').annotate(count=Count('pk')).order_by('current_level')

    # نرخ قبولی بر اساس سطح
    pass_rate_per_level = StudentGrade.objects.filter(
        exam_section__exam__status=Exam.Status.FINALIZED
    ).values(
        'exam_section__exam__class_group__course__level'
    ).annotate(
        avg_score=Avg('score'),
        student_count=Count('student', distinct=True)
    ).order_by('exam_section__exam__class_group__course__level')

    # ==================== ۷. جدول‌ها ====================
    pending_invoices = Invoice.objects.filter(status=Invoice.Status.PENDING).select_related('student__user', 'class_group__course').order_by('-created_at')[:10]
    paid_count = Invoice.objects.filter(status=Invoice.Status.PAID).count()
    pending_count = Invoice.objects.filter(status=Invoice.Status.PENDING).count()

    current_classes = Class.objects.filter(end_date__gte=today).select_related('course', 'teacher__user').order_by('start_date')
    if language_filter:
        current_classes = current_classes.filter(course__language__iexact=language_filter)
    if level_filter:
        current_classes = current_classes.filter(course__level__iexact=level_filter)
    current_classes = current_classes[:10]

    # ==================== آمار امتحانات برای جدول ====================
    exam_stats = StudentGrade.objects.filter(
        exam_section__exam__status=Exam.Status.FINALIZED
    ).values(
        'exam_section__exam__class_group__course__level'
    ).annotate(
        avg_score=Avg('score'),
        count_students=Count('student', distinct=True)
    ).order_by('exam_section__exam__class_group__course__level')

    # ==================== آرایه‌های آماده برای نمودار نرخ قبولی ====================
    pass_rate_labels = []
    pass_rate_data = []
    for item in pass_rate_per_level:
        pass_rate_labels.append(item['exam_section__exam__class_group__course__level'])
        pass_rate_data.append(round(item['avg_score'], 1))

    # ==================== ۹. Teacher Feedback (بازخورد اساتید) ====================
        # ==================== ۹. Teacher Feedback ====================
    teacher_feedbacks = ClassFeedback.objects.values(
        'class_group__teacher__user__first_name',
        'class_group__teacher__user__last_name',
        'class_group__teacher__department'
    ).annotate(
        avg_teaching=Avg('teaching_quality'),
        avg_communication=Avg('communication'),
        avg_punctuality=Avg('punctuality'),
        avg_engagement=Avg('engagement'),
        avg_overall=Avg('overall_satisfaction'),
        total_feedbacks=Count('id')
    ).order_by('class_group__teacher__user__last_name')

    # محاسبه آمار کلی
    total_feedbacks_count = sum(fb['total_feedbacks'] for fb in teacher_feedbacks) if teacher_feedbacks else 0
    overall_avg_rating = 0  # پیش‌فرض
    if teacher_feedbacks:
        overall_avg_rating = round(
            sum(fb['avg_overall'] for fb in teacher_feedbacks) / len(teacher_feedbacks), 1
        )
    context = {
        # فیلترها
        'date_range': date_range,
        'start_date': start_date_str if date_range == 'custom' else start_date.strftime('%Y-%m-%d'),
        'end_date': end_date_str if date_range == 'custom' else end_date.strftime('%Y-%m-%d'),
        'language_filter': language_filter,
        'level_filter': level_filter,
        'languages': Course.objects.values_list('language', flat=True).distinct().order_by('language'),
        'levels': ['A1', 'A2', 'B1', 'B2', 'C1', 'C2'],
        # KPI
        'total_revenue': total_revenue,
        'active_students': active_students,
        'active_classes_count': active_classes_count,
        'total_students': total_students_all,
        'total_classes': total_classes_all,
        'paid_count': paid_count,
        'pending_count': pending_count,
        'attendance_percent': attendance_percent,
        'completion_rate': completion_rate,
        'avg_score': round(avg_score, 1),
        'overall_pass_rate': overall_pass_rate,
        'pass_rate': pass_rate_per_level,
        # نمودارهای روند
        'monthly_revenue_labels': monthly_revenue_labels,
        'monthly_revenue_data': monthly_revenue_data,
        'enrollment_labels': enrollment_labels,
        'enrollment_data': enrollment_data,
        'weekly_att_labels': weekly_att_labels,
        'weekly_att_data': weekly_att_data,
        # نمودار توزیع
        'pie_language_labels': pie_language_labels,
        'pie_language_data': pie_language_data,
        'revenue_by_language': revenue_by_language,
        'level_distribution': level_distribution,
        # جدول‌ها
        'pending_invoices': pending_invoices,
        'current_classes': current_classes,
        'exam_stats': exam_stats,
        'pass_rate_labels': pass_rate_labels,
        'pass_rate_data': pass_rate_data,
        # Teacher Feedback
        'teacher_feedbacks': teacher_feedbacks,
        'total_feedbacks_count': total_feedbacks_count,
        'overall_avg_rating': overall_avg_rating,

    }
    return render(request, 'base/manager_reports.html', context)


@login_required(login_url='login')
def user_management(request):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position != Employee.Position.SENIOR_MANAGER
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    search_query = request.GET.get('q', '')
    
    users = UserAccount.objects.filter(
        is_superuser=False
    ).select_related('employee_profile').order_by('last_name', 'first_name')
    
    if search_query:
        users = users.filter(
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(employee_profile__position__icontains=search_query)
        )
    
    context = {
        'users': users,
        'search_query': search_query,
    }
    return render(request, 'base/user_management.html', context)


@login_required(login_url='login')
def finance_reports(request):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position != Employee.Position.SENIOR_MANAGER
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    search_query = request.GET.get('q', '')
    status_filter = request.GET.get('status', 'all')
    
    invoices = Invoice.objects.select_related(
        'student__user', 'class_group__course'
    ).order_by('-created_at')
    
    if status_filter in ['PAID', 'PENDING', 'CANCELED']:
        invoices = invoices.filter(status=status_filter)
    
    if search_query:
        invoices = invoices.filter(
            Q(reference_code__icontains=search_query) |
            Q(student__user__first_name__icontains=search_query) |
            Q(student__user__last_name__icontains=search_query) |
            Q(student__user__email__icontains=search_query)
        )
    
    total_revenue = invoices.filter(status=Invoice.Status.PAID).aggregate(Sum('amount'))['amount__sum'] or 0
    total_pending = invoices.filter(status=Invoice.Status.PENDING).aggregate(Sum('amount'))['amount__sum'] or 0
    paid_count = invoices.filter(status=Invoice.Status.PAID).count()
    pending_count = invoices.filter(status=Invoice.Status.PENDING).count()
    
    context = {
        'invoices': invoices,
        'search_query': search_query,
        'status_filter': status_filter,
        'total_revenue': total_revenue,
        'total_pending': total_pending,
        'paid_count': paid_count,
        'pending_count': pending_count,
    }
    return render(request, 'base/finance_reports.html', context)


@login_required(login_url='login')
def create_user(request):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position != Employee.Position.SENIOR_MANAGER
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip().lower()
        password = request.POST.get('password', '')
        position = request.POST.get('position', '')
        
        if not all([first_name, last_name, email, password, position]):
            messages.error(request, 'All fields are required.')
            return redirect('create_user')
        
        if UserAccount.objects.filter(email=email).exists():
            messages.error(request, 'A user with this email already exists.')
            return redirect('create_user')
        
        user = UserAccount.objects.create_user(
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            username=email,
            is_active=True
        )
        
        Employee.objects.create(
            user=user,
            position=position
        )
        
        messages.success(request, f'User {user.get_full_name()} created successfully.')
        return redirect('user_management')
    
    context = {
        'positions': Employee.Position.choices,
    }
    return render(request, 'base/create_user.html', context)


@login_required(login_url='login')
def manage_user(request, user_id):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position != Employee.Position.SENIOR_MANAGER
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    target_user = get_object_or_404(UserAccount, id=user_id)
    
    if target_user.is_superuser:
        messages.error(request, 'Cannot edit superuser accounts.')
        return redirect('user_management')
    
    if request.method == 'POST':
        new_position = request.POST.get('position', '')
        new_is_active = request.POST.get('is_active', 'on') == 'on'
        
        try:
            employee = target_user.employee_profile
            valid_positions = [choice[0] for choice in Employee.Position.choices]
            if new_position in valid_positions:
                employee.position = new_position
                employee.save()
        except Employee.DoesNotExist:
            pass
        
        target_user.is_active = new_is_active
        target_user.save()
        
        messages.success(request, f'User {target_user.get_full_name()} updated successfully.')
        return redirect('user_management')
    
    context = {
        'target_user': target_user,
        'positions': Employee.Position.choices,
    }
    return render(request, 'base/manage_user.html', context)


@login_required(login_url='login')
def deactivate_user(request, user_id):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position != Employee.Position.SENIOR_MANAGER
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    target_user = get_object_or_404(UserAccount, id=user_id)
    
    if target_user.is_superuser:
        messages.error(request, 'Cannot deactivate superuser accounts.')
        return redirect('user_management')
    
    if target_user == request.user:
        messages.error(request, 'You cannot deactivate your own account.')
        return redirect('user_management')
    
    target_user.is_active = False
    target_user.save()
    
    messages.success(request, f'User {target_user.get_full_name()} has been deactivated.')
    return redirect('user_management')

@login_required(login_url='login')
def create_class(request):
    # دسترسی: مدیر آموزشی، مدیر ارشد، یا ادمین
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        # خواندن داده‌ها از فرم
        title = request.POST.get('title', '').strip()
        course_id = request.POST.get('course')
        teacher_id = request.POST.get('teacher')
        tuition_fee = request.POST.get('tuition_fee', 0)
        capacity = request.POST.get('capacity', 10)
        class_type = request.POST.get('class_type', 'IN_PERSON')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        schedule = request.POST.get('schedule', '').strip()
        location = request.POST.get('location', '').strip()
        meeting_link = request.POST.get('meeting_link', '').strip()
        
        # روزهای هفته (از چک‌باکس‌ها)
        days_of_week = request.POST.getlist('days_of_week')  # لیستی از اعداد ۰ تا ۶
        days_of_week = [int(d) for d in days_of_week]
        
        # ساعت شروع و پایان
        start_time = request.POST.get('start_time') or None
        end_time = request.POST.get('end_time') or None
        
        # اعتبارسنجی ساده
        if not title or not course_id or not start_date or not end_date:
            messages.error(request, 'Please fill in all required fields.')
            return redirect('create_class')
        
        course = get_object_or_404(Course, pk=course_id)
        teacher = get_object_or_404(Employee, pk=teacher_id) if teacher_id else None
        
        # ساخت کلاس
        cls = Class.objects.create(
            title=title,
            course=course,
            teacher=teacher,
            tuition_fee=tuition_fee,
            capacity=capacity,
            class_type=class_type,
            start_date=start_date,
            end_date=end_date,
            schedule=schedule,
            location=location if class_type == 'IN_PERSON' else '',
            meeting_link=meeting_link if class_type == 'ONLINE' else '',
            day_of_week=days_of_week,
            start_time=start_time,
            end_time=end_time,
        )
        
        messages.success(request, f'Class "{cls.title}" created successfully with code {cls.class_code}.')
        return redirect('manage_classes')
    
    # GET: نمایش فرم
    courses = Course.objects.all().order_by('language', 'level')
    teachers = Employee.objects.filter(position=Employee.Position.TEACHER).select_related('user').order_by('user__last_name')
    
    context = {
        'courses': courses,
        'teachers': teachers,
        'days_choices': [
            (0, 'Mon'), (1, 'Tue'), (2, 'Wed'), (3, 'Thu'),
            (4, 'Fri'), (5, 'Sat'), (6, 'Sun')
        ],
    }
    return render(request, 'base/create_class.html', context)


@login_required(login_url='login')
def edit_class(request, class_id):
    # دسترسی: مدیر آموزشی، مدیر ارشد، یا ادمین
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER,
            Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    cls = get_object_or_404(Class, id=class_id)
    
    if request.method == 'POST':
        # خواندن داده‌ها از فرم
        title = request.POST.get('title', '').strip()
        course_id = request.POST.get('course')
        teacher_id = request.POST.get('teacher')
        tuition_fee = request.POST.get('tuition_fee', 0)
        capacity = request.POST.get('capacity', 10)
        class_type = request.POST.get('class_type', 'IN_PERSON')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        schedule = request.POST.get('schedule', '').strip()
        location = request.POST.get('location', '').strip()
        meeting_link = request.POST.get('meeting_link', '').strip()
        
        # روزهای هفته
        days_of_week = request.POST.getlist('days_of_week')
        days_of_week = [int(d) for d in days_of_week]
        
        # ساعت شروع و پایان
        start_time = request.POST.get('start_time') or None
        end_time = request.POST.get('end_time') or None
        
        if not title or not course_id or not start_date or not end_date:
            messages.error(request, 'Please fill in all required fields.')
            return redirect('edit_class', class_id=cls.id)
        
        course = get_object_or_404(Course, pk=course_id)
        teacher = get_object_or_404(Employee, pk=teacher_id) if teacher_id else None
        
        # به‌روزرسانی کلاس
        cls.title = title
        cls.course = course
        cls.teacher = teacher
        cls.tuition_fee = tuition_fee
        cls.capacity = capacity
        cls.class_type = class_type
        cls.start_date = start_date
        cls.end_date = end_date
        cls.schedule = schedule
        cls.location = location if class_type == 'IN_PERSON' else ''
        cls.meeting_link = meeting_link if class_type == 'ONLINE' else ''
        cls.day_of_week = days_of_week
        cls.start_time = start_time
        cls.end_time = end_time
        cls.save()
        
        messages.success(request, f'Class "{cls.title}" updated successfully.')
        return redirect('manage_classes')
    
    # GET: نمایش فرم با داده‌های قبلی
    courses = Course.objects.all().order_by('language', 'level')
    teachers = Employee.objects.filter(position=Employee.Position.TEACHER).select_related('user').order_by('user__last_name')
    
    context = {
        'class': cls,
        'courses': courses,
        'teachers': teachers,
        'days_choices': [
            (0, 'Mon'), (1, 'Tue'), (2, 'Wed'), (3, 'Thu'),
            (4, 'Fri'), (5, 'Sat'), (6, 'Sun')
        ],
    }
    return render(request, 'base/edit_class.html', context)


@login_required(login_url='login')
def placement_test_settings(request):
    # فقط مدیر ارشد
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position != Employee.Position.SENIOR_MANAGER
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    settings = PlacementTestSettings.load()
    
    if request.method == 'POST':
        test_fee = request.POST.get('test_fee', 0)
        settings.test_fee = test_fee
        settings.save()
        messages.success(request, f'Placement test fee updated to {test_fee} T.')
        return redirect('placement_test_settings')
    
    context = {'settings': settings}
    return render(request, 'base/placement_test_settings.html', context)

@login_required(login_url='login')
def manage_placement_requests(request):
    # دسترسی: مدیر آموزشی، مدیر ارشد
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    status_filter = request.GET.get('status', '')
    requests = PlacementTestRequest.objects.select_related('student__user').order_by('-created_at')
    pending_count = requests.filter(status=PlacementTestRequest.Status.PENDING).count()
    approved_count = requests.filter(status=PlacementTestRequest.Status.APPROVED).count()
    rejected_count = requests.filter(status=PlacementTestRequest.Status.REJECTED).count()
    paid_requests_count = requests.filter(payment_status=Enrollment.PaymentStatus.PAID).count()
    
    if status_filter:
        requests = requests.filter(status=status_filter)
        
    context = {
        'requests': requests,
        'status_filter': status_filter,
        'status_choices': PlacementTestRequest.Status.choices,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'paid_requests_count': paid_requests_count,
    }
    return render(request, 'base/manage_placement_requests.html', context)

@login_required(login_url='login')
def review_placement_request(request, request_id):
    # دسترسی مثل بالا
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')
    
    placement_request = get_object_or_404(PlacementTestRequest, id=request_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')  # 'approve' or 'reject'
        test_date = request.POST.get('test_date', '')
        test_time = request.POST.get('test_time', '')
        
        if action == 'approve':
            placement_request.status = PlacementTestRequest.Status.APPROVED
            if test_date:
                placement_request.test_date = date.fromisoformat(test_date)
            if test_time:
                placement_request.test_time = test_time
            placement_request.save()
            messages.success(request, 'Request approved and test scheduled.')
        elif action == 'reject':
            placement_request.status = PlacementTestRequest.Status.REJECTED
            placement_request.save()
            messages.warning(request, 'Request rejected.')
        
        return redirect('manage_placement_requests')
    
    context = {'placement_request': placement_request}
    return render(request, 'base/review_placement_request.html', context)


@login_required(login_url='login')
def manage_withdrawal_requests(request):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    status_filter = request.GET.get('status', '')
    requests = WithdrawalRequest.objects.select_related('student__user', 'enrollment__enrolled_class').order_by('-created_at')
    if status_filter:
        requests = requests.filter(status=status_filter)

    context = {
        'requests': requests,
        'status_filter': status_filter,
        'status_choices': WithdrawalRequest.Status.choices,
    }
    return render(request, 'base/manage_withdrawal_requests.html', context)

@login_required(login_url='login')
def manage_transfer_requests(request):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    status_filter = request.GET.get('status', '')
    requests = TransferRequest.objects.select_related(
        'student__user', 'from_enrollment__enrolled_class', 'to_class__course'
    ).order_by('-created_at')
    if status_filter:
        requests = requests.filter(status=status_filter)

    pending_count = requests.filter(status=TransferRequest.Status.PENDING).count() if not status_filter else None

    context = {
        'requests': requests,
        'status_filter': status_filter,
        'status_choices': TransferRequest.Status.choices,
        'pending_count': pending_count,
    }
    return render(request, 'base/manage_transfer_requests.html', context)

@login_required(login_url='login')
def approve_transfer_request(request, request_id):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    transfer_req = get_object_or_404(TransferRequest, id=request_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            # تراکنش اتمی: کنسل کردن قدیم + ثبت‌نام جدید + مالی
            from django.db import transaction
            try:
                with transaction.atomic():
                    # ۱. کنسل کردن ثبت‌نام قبلی
                    from_enrollment = transfer_req.from_enrollment
                    from_enrollment.status = Enrollment.EnrollmentStatus.CANCELED
                    from_enrollment.save()

                    # ۲. ایجاد ثبت‌نام جدید
                    to_class = transfer_req.to_class
                    new_enrollment = Enrollment.objects.create(
                        student=transfer_req.student,
                        enrolled_class=to_class,
                        payment_status=Enrollment.PaymentStatus.PAID
                    )

                    # ۳. مدیریت مالی
                    old_fee = from_enrollment.enrolled_class.tuition_fee
                    new_fee = to_class.tuition_fee
                    diff = new_fee - old_fee
                    if diff > 0:
                        tax_settings = TaxSettings.load()
                        tax_amount = (diff * tax_settings.tax_percent) / 100
                        
                        Invoice.objects.create(
                            student=transfer_req.student,
                            class_group=to_class,
                            amount=diff,
                            tax_amount=tax_amount,
                            status=Invoice.Status.PENDING,
                            reference_code=f"TRANS-{uuid.uuid4().hex[:8].upper()}"
                        )
                    elif diff < 0:
                        tax_settings = TaxSettings.load()
                        tax_amount = (abs(diff) * tax_settings.tax_percent) / 100
                        
                        Invoice.objects.create(
                            student=transfer_req.student,
                            class_group=None,
                            amount=abs(diff),
                            tax_amount=tax_amount,
                            status=Invoice.Status.PAID,
                            reference_code=f"REFUND-{uuid.uuid4().hex[:8].upper()}"
                        )

                    # ۴. به‌روزرسانی درخواست
                    transfer_req.status = TransferRequest.Status.APPROVED
                    transfer_req.reviewed_by = request.user.employee_profile
                    transfer_req.save()

                messages.success(request, 'Transfer request approved and processed.')
            except Exception as e:
                messages.error(request, f'Transfer failed: {str(e)}')
        elif action == 'reject':
            transfer_req.status = TransferRequest.Status.REJECTED
            transfer_req.reviewed_by = request.user.employee_profile
            transfer_req.save()
            messages.success(request, 'Transfer request rejected.')
        return redirect('manage_transfer_requests')

    context = {'transfer_req': transfer_req}
    return render(request, 'base/approve_transfer_request.html', context)

@login_required(login_url='login')
def review_withdrawal_request(request, request_id):
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    withdrawal = get_object_or_404(WithdrawalRequest, id=request_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            withdrawal.status = WithdrawalRequest.Status.APPROVED
            withdrawal.reviewed_by = request.user.employee_profile
            withdrawal.save()
            # کنسل کردن enrollment
            enrollment = withdrawal.enrollment
            enrollment.status = Enrollment.EnrollmentStatus.CANCELED
            enrollment.save()
            # (اختیاری) می‌تونی یه Invoice cancellation هم اینجا اضافه کنی
            messages.success(request, 'Withdrawal approved. Enrollment has been canceled.')
        elif action == 'reject':
            withdrawal.status = WithdrawalRequest.Status.REJECTED
            withdrawal.reviewed_by = request.user.employee_profile
            withdrawal.save()
            messages.success(request, 'Withdrawal request rejected.')
        return redirect('manage_withdrawal_requests')

    context = {'withdrawal': withdrawal}
    return render(request, 'base/review_withdrawal_request.html', context)


@login_required(login_url='login')
def submit_feedback(request, class_id):
    cls = get_object_or_404(Class, id=class_id)
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, 'Only students can submit feedback.')
        return redirect('dashboard')
    
    # چک اینکه آیا امتحان نهایی شده است یا نه
    if not hasattr(cls, 'exam') or cls.exam.status != Exam.Status.FINALIZED:
        messages.error(request, 'Feedback can only be submitted after the exam is finalized.')
        return redirect('student_scores')
    
    # چک اینکه قبلاً بازخورد ثبت نشده باشد
    if ClassFeedback.objects.filter(student=student, class_group=cls).exists():
        messages.warning(request, 'You have already submitted feedback for this class.')
        return redirect('student_scores')
    
    if request.method == 'POST':
        teaching_quality = request.POST.get('teaching_quality')
        communication = request.POST.get('communication')
        punctuality = request.POST.get('punctuality')
        engagement = request.POST.get('engagement')
        overall_satisfaction = request.POST.get('overall_satisfaction')
        comments = request.POST.get('comments', '')
        
        # اعتبارسنجی ساده
        if not all([teaching_quality, communication, punctuality, engagement, overall_satisfaction]):
            messages.error(request, 'Please answer all rating questions.')
            return redirect('submit_feedback', class_id=cls.id)
        
        ClassFeedback.objects.create(
            student=student,
            class_group=cls,
            teaching_quality=int(teaching_quality),
            communication=int(communication),
            punctuality=int(punctuality),
            engagement=int(engagement),
            overall_satisfaction=int(overall_satisfaction),
            comments=comments
        )
        messages.success(request, 'Thank you for your feedback! You can now view your scores.')
        return redirect('student_scores')
    
    context = {
        'class': cls,
    }
    return render(request, 'base/submit_feedback.html', context)


@login_required(login_url='login')
def request_withdrawal(request, enrollment_id):
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, 'Only students can request withdrawal.')
        return redirect('dashboard')

    enrollment = get_object_or_404(Enrollment, id=enrollment_id, student=student)

    if enrollment.status != Enrollment.EnrollmentStatus.ACTIVE:
        messages.error(request, 'This enrollment is not active.')
        return redirect('dashboard')
    
    if enrollment.enrolled_class.end_date and enrollment.enrolled_class.end_date < date.today():
        messages.error(request, 'This class has already ended. Withdrawal is not possible.')
        return redirect('dashboard')

    if WithdrawalRequest.objects.filter(enrollment=enrollment, status=WithdrawalRequest.Status.PENDING).exists():
        messages.warning(request, 'You already have a pending withdrawal request for this class.')
        return redirect('dashboard')

    if request.method == 'POST':
        reason = request.POST.get('reason', '')
        WithdrawalRequest.objects.create(
            student=student,
            enrollment=enrollment,
            reason=reason
        )
        messages.success(request, 'Your withdrawal request has been submitted and is pending approval.')
        return redirect('dashboard')

    context = {'enrollment': enrollment}
    return render(request, 'base/request_withdrawal.html', context)

@login_required(login_url='login')
def transfer_student(request, student_id):
    # ۱. چک دسترسی: EDUCATION_MANAGER, SENIOR_MANAGER
    if not request.user.is_staff and (
        not hasattr(request.user, 'employee_profile') or
        request.user.employee_profile.position not in [
            Employee.Position.EDUCATION_MANAGER, Employee.Position.SENIOR_MANAGER
        ]
    ):
        messages.error(request, 'You do not have permission.')
        return redirect('dashboard')

    student = get_object_or_404(Student.objects.select_related('user'), pk=student_id)
    
    # لیست کلاس‌های فعال این دانشجو (برای انتخاب مبدأ)
    active_enrollments = Enrollment.objects.filter(
        student=student,
        status=Enrollment.EnrollmentStatus.ACTIVE,
        enrolled_class__end_date__gte=date.today()
    ).select_related('enrolled_class__course', 'enrolled_class__teacher__user')
    
    # لیست کلاس‌های موجود برای انتقال (مقصد) - کلاس‌های جاری که دانشجو در آن‌ها نیست
    available_classes = Class.objects.filter(
        end_date__gte=date.today()
    ).exclude(
        id__in=active_enrollments.values_list('enrolled_class_id', flat=True)
    ).select_related('course', 'teacher__user').order_by('start_date')

    if request.method == 'POST':
        from_enrollment_id = request.POST.get('from_enrollment')
        to_class_id = request.POST.get('to_class')
        
        if not from_enrollment_id or not to_class_id:
            messages.error(request, 'Please select both source and destination classes.')
            return redirect('transfer_student', student_id=student_id)
        
        from_enrollment = get_object_or_404(Enrollment, id=from_enrollment_id, student=student)
        to_class = get_object_or_404(Class, id=to_class_id)
        
        # اعتبارسنجی‌ها
        if from_enrollment.status != Enrollment.EnrollmentStatus.ACTIVE:
            messages.error(request, 'The selected enrollment is not active.')
            return redirect('transfer_student', student_id=student_id)
        
        if to_class.end_date < date.today():
            messages.error(request, 'The destination class has already ended.')
            return redirect('transfer_student', student_id=student_id)
        
        if to_class.enrollments.count() >= to_class.capacity:
            messages.error(request, 'The destination class is full.')
            return redirect('transfer_student', student_id=student_id)
        
        if Enrollment.objects.filter(student=student, enrolled_class=to_class).exists():
            messages.error(request, 'Student is already enrolled in the destination class.')
            return redirect('transfer_student', student_id=student_id)
        
        # مدیریت شهریه
        old_fee = from_enrollment.enrolled_class.tuition_fee
        new_fee = to_class.tuition_fee
        fee_difference = new_fee - old_fee
        
        # تراکنش اتمی
        from django.db import transaction
        try:
            with transaction.atomic():
                # کنسل کردن ثبت‌نام قبلی
                from_enrollment.status = Enrollment.EnrollmentStatus.CANCELED
                from_enrollment.save()
                
                # ایجاد ثبت‌نام جدید
                new_enrollment = Enrollment.objects.create(
                    student=student,
                    enrolled_class=to_class,
                    payment_status=Enrollment.PaymentStatus.PAID  # فرض می‌کنیم پرداخت شده
                )
                
                # رسیدگی مالی
                if fee_difference > 0:
                    tax_settings = TaxSettings.load()
                    tax_amount = (fee_difference * tax_settings.tax_percent) / 100
                    
                    # ایجاد فاکتور برای مابه‌التفاوت
                    Invoice.objects.create(
                        student=student,
                        class_group=to_class,
                        amount=fee_difference,
                        tax_amount=tax_amount,
                        status=Invoice.Status.PENDING,
                        reference_code=f"TRANS-{uuid.uuid4().hex[:8].upper()}"
                    )
                    messages.success(request, f'Transfer successful. Additional payment of {fee_difference} T is required.')
                elif fee_difference < 0:
                    tax_settings = TaxSettings.load()
                    tax_amount = (abs(fee_difference) * tax_settings.tax_percent) / 100
                    # می‌تونیم اعتبار ثبت کنیم یا فاکتور برگشتی
                    Invoice.objects.create(
                        student=student,
                        class_group=None,
                        amount=abs(fee_difference),
                        tax_amount=tax_amount,
                        status=Invoice.Status.PAID,  # به عنوان اعتبار در نظر گرفته شود
                        reference_code=f"REFUND-{uuid.uuid4().hex[:8].upper()}"
                    )
                    messages.success(request, f'Transfer successful. A credit of {abs(fee_difference)} T has been issued.')
                else:
                    messages.success(request, 'Transfer successful.')
                
                return redirect('student_detail', student_id=student_id)
        except Exception as e:
            messages.error(request, f'Transfer failed: {str(e)}')
            return redirect('transfer_student', student_id=student_id)

    context = {
        'student': student,
        'active_enrollments': active_enrollments,
        'available_classes': available_classes,
    }
    return render(request, 'base/transfer_student.html', context)